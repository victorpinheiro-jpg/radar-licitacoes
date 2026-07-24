import streamlit as st
import pandas as pd
import requests
from datetime import datetime, timedelta
import time
import os
import io
import re
import xml.etree.ElementTree as ET
import urllib.parse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- 1. CONFIGURAÇÃO VISUAL E IDENTIDADE ASA ---
st.set_page_config(page_title="ASA | Radar de Infraestrutura", page_icon="⚖️", layout="wide")

st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .block-container { padding-top: 2rem; }
    .stButton>button { width: 100%; border-radius: 8px; font-weight: 500; background-color: #6a9094; color: white; height: 3em; border: none; transition: all 0.3s ease; }
    .stButton>button:hover { background-color: #55787c; color: white; }
    div[data-testid="stMetricValue"] { color: #436468; font-size: 2.2rem; font-weight: 700; }
    h1, h2, h3, h4, h5, h6 { color: #436468 !important; font-weight: 600; }
    hr { border-bottom-color: #6a9094 !important; opacity: 0.3; }
    </style>
""", unsafe_allow_html=True)

if 'licitacoes_salvas' not in st.session_state: 
    st.session_state['licitacoes_salvas'] = pd.DataFrame()
if 'resultados_busca' not in st.session_state: 
    st.session_state['resultados_busca'] = pd.DataFrame()
if 'busca_realizada' not in st.session_state: 
    st.session_state['busca_realizada'] = False

MAPA_MODALIDADES = {
    "Leilão": 1, "Diálogo Competitivo": 2, "Concurso": 3, 
    "Concorrência": 4, "Pregão Eletrônico": 6
}
LISTA_UFS = ["AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"]

# --- FUNÇÕES DE APOIO E ESTILIZAÇÃO ---
def calcular_dias_restantes(data_sessao_str):
    if not data_sessao_str or data_sessao_str == "Verificar Edital": 
        return "N/A"
    try:
        dt_sessao = datetime.strptime(data_sessao_str, '%d/%m/%Y %H:%M')
        hoje = datetime.now()
        delta = (dt_sessao.date() - hoje.date()).days
        if delta < 0: 
            return "Sessão Encerrada"
        elif delta == 0: 
            return "🚨 É HOJE!"
        else: 
            return f"Faltam {delta} dias"
    except: 
        return "N/A"

def aplicar_estilo_excel(writer, df, sheet_name):
    worksheet = writer.sheets[sheet_name]
    
    cor_asa = PatternFill(start_color="436468", end_color="436468", fill_type="solid")
    cor_fundo_claro = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    cor_suspensa = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") 
    cor_morta = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cor_alvo = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    fonte_branca = Font(color="FFFFFF", bold=True)
    fonte_normal = Font(name="Calibri", size=11)
    fonte_urgente = Font(name="Calibri", size=11, color="FF0000", bold=True)
    
    borda_fina = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), 
                        top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))
    
    worksheet.row_dimensions[1].height = 30
    col_indices = {cell.value: idx + 1 for idx, cell in enumerate(worksheet[1])}
    status_col_idx = col_indices.get("Status/Fase")
    dias_col_idx = col_indices.get("Dias Restantes")

    for cell in worksheet[1]:
        cell.fill = cor_asa
        cell.font = fonte_branca
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda_fina
        
    for col in worksheet.columns:
        col_letter = col[0].column_letter
        col_name = col[0].value
        if col_name in ["Objeto", "Manchete / Título"]: 
            worksheet.column_dimensions[col_letter].width = 65
        elif col_name in ["Órgão", "Empresa Vencedora (Alvo)"]: 
            worksheet.column_dimensions[col_letter].width = 40
        elif col_name in ["Link", "Link da Notícia", "Anotações Equipe"]: 
            worksheet.column_dimensions[col_letter].width = 45
        elif col_name in ["Status/Fase", "Última Atualização", "Data da Sessão", "Dias Restantes", "Data da Notícia", "Fonte"]: 
            worksheet.column_dimensions[col_letter].width = 20
        elif col_name in ["Valor Estimado", "Valor Arrematado"]: 
            worksheet.column_dimensions[col_letter].width = 22
        else: 
            worksheet.column_dimensions[col_letter].width = 18
            
    for row_idx in range(2, worksheet.max_row + 1):
        status_val = str(worksheet.cell(row=row_idx, column=status_col_idx).value).lower() if status_col_idx else ""
        dias_val = str(worksheet.cell(row=row_idx, column=dias_col_idx).value) if dias_col_idx else ""
        
        fundo_linha = PatternFill(fill_type=None)
        if row_idx % 2 == 0: 
            fundo_linha = cor_fundo_claro
            
        if "suspen" in status_val: 
            fundo_linha = cor_suspensa
        elif any(x in status_val for x in ["homolog", "revogad", "cancelad", "fracassad", "desert"]): 
            fundo_linha = cor_morta
            
        if "Radar de Prospecção" in sheet_name or "Vencedores" in sheet_name: 
            if row_idx % 2 == 0:
                fundo_linha = cor_alvo 
            else:
                fundo_linha = PatternFill(fill_type=None)

        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            col_name = worksheet.cell(row=1, column=col_idx).value
            
            cell.fill = fundo_linha
            cell.border = borda_fina
            cell.font = fonte_normal
            
            if col_name == "Dias Restantes":
                if "HOJE" in dias_val: 
                    cell.font = fonte_urgente
                elif "Faltam" in dias_val:
                    try:
                        if int(re.search(r'\d+', dias_val).group()) <= 5: 
                            cell.font = fonte_urgente 
                    except: 
                        pass
                        
            if col_name in ["Objeto", "Anotações Equipe", "Empresa Vencedora (Alvo)", "Manchete / Título"]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                texto_len = len(str(cell.value)) if cell.value else 0
                linhas_estimadas = max(1, (texto_len // 60) + 1)
                if worksheet.row_dimensions[row_idx].height is None or worksheet.row_dimensions[row_idx].height < linhas_estimadas * 15:
                    worksheet.row_dimensions[row_idx].height = linhas_estimadas * 15
            else: 
                cell.alignment = Alignment(vertical="top")
                
            if col_name in ["Valor Estimado", "Valor Arrematado"] and isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'
                
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions

# --- 2. LÓGICA DE MANIPULAÇÃO DO PIPELINE ÚNICO ---
def load_master_excel(file_buffer):
    try:
        xls = pd.read_excel(file_buffer, sheet_name=None)
        df = pd.concat([v for k, v in xls.items() if not v.empty], ignore_index=True)
        if df.empty:
            return pd.DataFrame()
        
        cols_defaults = {
            "Empresa Vencedora (Alvo)": "Ainda sem vencedor publicado",
            "CNPJ do Alvo": "-",
            "Valor Arrematado": 0.0
        }
        for col, val in cols_defaults.items():
            if col not in df.columns:
                df[col] = val
            df[col] = df[col].fillna(val)
            
        return df
    except Exception as e:
        return pd.DataFrame()

def gerar_excel_pipeline(df_master):
    df_vencedores = df_master[df_master["Empresa Vencedora (Alvo)"] != "Ainda sem vencedor publicado"].copy()
    df_transito = df_master[df_master["Empresa Vencedora (Alvo)"] == "Ainda sem vencedor publicado"].copy()
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as w:
        if not df_transito.empty:
            df_transito.to_excel(w, index=False, sheet_name='Em Trânsito')
            aplicar_estilo_excel(w, df_transito, 'Em Trânsito')
            
        if not df_vencedores.empty:
            df_vencedores.to_excel(w, index=False, sheet_name='Vencedores (Alvos)')
            aplicar_estilo_excel(w, df_vencedores, 'Vencedores (Alvos)')
            
        if df_vencedores.empty and df_transito.empty:
            pd.DataFrame().to_excel(w, index=False, sheet_name='Vazio')
            
    return buffer, len(df_transito), len(df_vencedores)

# --- 3. MOTORES DE BUSCA (MODO APRESENTAÇÃO ATIVADO) ---
# Foi removido o "@st.cache_data" para o robô NÃO gravar erros na memória
def buscar_licitacoes_periodo(data_inicio, data_fim, modalidades_selecionadas):
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}
    todos_resultados = []
    erros = []
    
    if not modalidades_selecionadas: 
        modalidades_selecionadas = ["Concorrência", "Leilão", "Diálogo Competitivo"]
        
    chunks = []
    atual = data_inicio
    while atual <= data_fim:
        proximo = min(atual + timedelta(days=7), data_fim)
        chunks.append((atual, proximo))
        atual = proximo + timedelta(days=1)
        
    # 🌟 MODO APRESENTAÇÃO: Barra de progresso visual
    progresso_busca = st.progress(0, text="Conectando aos servidores do PNCP...")
    total_passos = len(modalidades_selecionadas) * len(chunks)
    passo_atual = 0
        
    for modalidade in modalidades_selecionadas:
        codigo = MAPA_MODALIDADES.get(modalidade)
        if not codigo: 
            continue
            
        for inicio_chunk, fim_chunk in chunks:
            str_inicio = inicio_chunk.strftime("%Y%m%d")
            str_fim = fim_chunk.strftime("%Y%m%d")
            
            pagina = 1
            tem_mais_paginas = True
            
            while tem_mais_paginas:
                url = f"https://pncp.gov.br/api/consulta/v1/contratacoes/publicacao?dataInicial={str_inicio}&dataFinal={str_fim}&codigoModalidadeContratacao={codigo}&pagina={pagina}&tamanhoPagina=50"
                
                sucesso = False
                tentativas = 0
                
                # Falha rápido (2 tentativas) com timeout curto (10s) para não congelar a tela
                while not sucesso and tentativas < 2:
                    try:
                        response = requests.get(url, headers=headers, timeout=10)
                        if response.status_code == 200:
                            dados_pagina = response.json().get("data", [])
                            todos_resultados.extend(dados_pagina)
                            sucesso = True
                            
                            if len(dados_pagina) < 50:
                                tem_mais_paginas = False
                            else:
                                pagina += 1 
                                
                        elif response.status_code == 429: 
                            time.sleep(2)
                            tentativas += 1
                        else: 
                            tentativas += 1
                            time.sleep(1) 
                    except: 
                        tentativas += 1
                        time.sleep(1)
                
                if not sucesso: 
                    erros.append(f"{modalidade} ({str_inicio})")
                    tem_mais_paginas = False 
            
            # Atualiza a barrinha verde na tela para quem estiver assistindo
            passo_atual += 1
            progresso_busca.progress(passo_atual / total_passos, text=f"Varrendo {modalidade}: {inicio_chunk.strftime('%d/%m/%Y')}...")
            
    # Esconde a barra quando terminar
    progresso_busca.empty() 
    return todos_resultados, list(set(erros))

def filtrar_dados(licitacoes, palavras_chave, valor_min, valor_max, estados_selecionados):
    if not licitacoes: 
        return pd.DataFrame()
        
    resultados = []
    lista_palavras = [p.strip().lower() for p in palavras_chave.split(',')] if palavras_chave else []
    ids_adicionados = set()
    
    for lic in licitacoes:
        id_unico = lic.get("id") or lic.get("linkSistemaOrigem")
        if id_unico in ids_adicionados: 
            continue
        
        uf_licitacao = lic.get("unidadeOrgao", {}).get("ufSigla", "N/A")
        if estados_selecionados and (uf_licitacao not in estados_selecionados): 
            continue
            
        objeto = lic.get("objetoCompra") or lic.get("sinteseObjeto") or lic.get("objeto") or "Descrição indisponível"
        objeto_str = str(objeto).lower()
        valor = lic.get("valorTotalEstimado") or 0.0 
        
        passou_palavra = any(p in objeto_str for p in lista_palavras) if (lista_palavras and lista_palavras[0] != "") else True
        if passou_palavra and (valor_min <= valor <= valor_max):
            numero_compra = lic.get("numeroCompra", "")
            ano_compra = lic.get("anoCompra", "")
            sequencial_compra = lic.get("sequencialCompra", "")
            cnpj = lic.get("orgaoEntidade", {}).get("cnpj", "")
            
            if cnpj and ano_compra and str(sequencial_compra) != "": 
                link_final = f"https://pncp.gov.br/app/editais/{cnpj}/{ano_compra}/{sequencial_compra}"
            elif lic.get("linkSistemaOrigem"):
                link_original = str(lic.get("linkSistemaOrigem", ""))
                link_final = link_original if link_original.startswith("http") else "https://" + link_original
            else: 
                link_final = "https://pncp.gov.br"

            identificacao = f"{numero_compra}/{ano_compra}" if numero_compra and ano_compra else str(id_unico)
            fase = lic.get("situacaoCompraNome", "Não informada")
            
            dt_att = lic.get("dataAtualizacaoPncp")
            data_atualizacao_str = datetime.fromisoformat(dt_att[:19]).strftime('%d/%m/%Y %H:%M') if dt_att else "Não informada"

            dt_sessao = lic.get("dataAberturaProposta")
            data_sessao_str = datetime.fromisoformat(dt_sessao[:19]).strftime('%d/%m/%Y %H:%M') if dt_sessao else "Verificar Edital"

            resultados.append({
                "Identificação": identificacao, 
                "Status/Fase": fase, 
                "Dias Restantes": calcular_dias_restantes(data_sessao_str),
                "Data da Sessão": data_sessao_str, 
                "Última Atualização": data_atualizacao_str, 
                "Anotações Equipe": "", 
                "UF": uf_licitacao, 
                "Órgão": lic.get("orgaoEntidade", {}).get("razaoSocial", "N/A"),
                "Modalidade": lic.get("modalidadeNome", "N/A"), 
                "Objeto": objeto, 
                "Valor Estimado": valor,
                "Data Publicação": lic.get("dataPublicacaoPncp")[:10] if lic.get("dataPublicacaoPncp") else "", 
                "Link": link_final
            })
            ids_adicionados.add(id_unico)
            
    return pd.DataFrame(resultados)

# --- WORKERS DE PROCESSAMENTO PARALELO ---
def worker_rastrear(index, row):
    link = str(row["Link"]).strip()
    resultado = {
        "index": index, 
        "Status/Fase": row.get("Status/Fase", ""), 
        "Última Atualização": row.get("Última Atualização", ""), 
        "Data da Sessão": row.get("Data da Sessão", ""), 
        "Dias Restantes": row.get("Dias Restantes", "")
    }
    match = re.search(r"editais/(\d+)/(\d+)/(\d+)", link)
    if match:
        cnpj, ano, seq = match.groups()
        try:
            resp = requests.get(f"https://pncp.gov.br/api/pncp/v1/orgaos/{cnpj}/compras/{ano}/{seq}", headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
            if resp.status_code == 200:
                dados = resp.json()
                resultado["Status/Fase"] = dados.get("situacaoCompraNome", resultado["Status/Fase"])
                if dados.get("dataAtualizacaoPncp"): 
                    resultado["Última Atualização"] = datetime.fromisoformat(dados["dataAtualizacaoPncp"][:19]).strftime('%d/%m/%Y %H:%M')
                if dados.get("dataAberturaProposta"):
                    sessao = datetime.fromisoformat(dados["dataAberturaProposta"][:19]).strftime('%d/%m/%Y %H:%M')
                    resultado["Data da Sessão"] = sessao
                    resultado["Dias Restantes"] = calcular_dias_restantes(sessao)
        except: 
            pass
    return resultado

def worker_prospeccao(row):
    alvo = row.to_dict() 
    link = str(alvo.get("Link", "")).strip()
    
    alvo["Empresa Vencedora (Alvo)"] = "Ainda sem vencedor publicado"
    alvo["CNPJ do Alvo"] = "-"
    alvo["Valor Arrematado"] = 0.0 
    
    match = re.search(r"editais/(\d+)/(\d+)/(\d+)", link)
    if match:
        cnpj, ano, seq = match.groups()
        encontrou = False
        
        try: 
            r_c = requests.get(f"https://pncp.gov.br/api/pncp/v1/orgaos/{cnpj}/compras/{ano}/{seq}/contratos", headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
            if r_c.status_code == 200:
                lista = r_c.json() if isinstance(r_c.json(), list) else r_c.json().get("data", [])
                if lista:
                    alvo["Empresa Vencedora (Alvo)"] = lista[0].get("nomeRazaoSocialFornecedor", alvo["Empresa Vencedora (Alvo)"])
                    alvo["CNPJ do Alvo"] = lista[0].get("niFornecedor", alvo["CNPJ do Alvo"])
                    val = lista[0].get("valorInicial") or lista[0].get("valorGlobal")
                    if val: 
                        alvo["Valor Arrematado"] = float(val)
                    encontrou = True
        except: 
            pass
        
        if not encontrou:
            try: 
                r_r = requests.get(f"https://pncp.gov.br/api/pncp/v1/orgaos/{cnpj}/compras/{ano}/{seq}/itens/1/resultados", headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
                if r_r.status_code == 200:
                    lista = r_r.json() if isinstance(r_r.json(), list) else r_r.json().get("data", [])
                    if lista:
                        alvo["Empresa Vencedora (Alvo)"] = lista[0].get("nomeRazaoSocialFornecedor", alvo["Empresa Vencedora (Alvo)"])
                        alvo["CNPJ do Alvo"] = lista[0].get("niFornecedor", alvo["CNPJ do Alvo"])
                        val = lista[0].get("valorTotalHomologado")
                        if val: 
                            alvo["Valor Arrematado"] = float(val)
            except: 
                pass
    return alvo

def worker_rss(fase, tema, periodo_dias):
    query_str = f'"{fase}" AND "{tema}" when:{periodo_dias}d'
    query = urllib.parse.quote(query_str)
    url = f"https://news.google.com/rss/search?q={query}&hl=pt-BR&gl=BR&ceid=BR:pt-419"
    res = []
    
    termos_obrigatorios = [
        "licita", "concess", "ppp", "público-privada", "saneamento", "rodovia", 
        "pedágio", "leilão", "leilao", "privatiza", "pública", "pmi", "edital", 
        "infraestrutura", "obras", "desestatização", "manifestação de interesse", "b3"
    ]
    
    try:
        resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
        if resp.status_code == 200:
            root = ET.fromstring(resp.content)
            for item in root.findall('.//item'):
                titulo = item.find('title').text if item.find('title') is not None else ""
                titulo_lower = titulo.lower()
                
                if not any(termo in titulo_lower for termo in termos_obrigatorios):
                    continue 

                data_pub = item.find('pubDate').text if item.find('pubDate') is not None else ""
                try: 
                    dt_str = datetime.strptime(data_pub, "%a, %d %b %Y %H:%M:%S %Z").strftime("%d/%m/%Y %H:%M")
                except: 
                    dt_str = data_pub
                
                link = item.find('link').text if item.find('link') is not None else ""
                fonte = item.find('source').text if item.find('source') is not None else "Google News"
                
                res.append({
                    "Data da Notícia": dt_str,
                    "Fonte": fonte,
                    "Manchete / Título": titulo,
                    "Link da Notícia": link,
                    "Anotações Equipe": ""
                })
    except: 
        pass
        
    return pd.DataFrame(res)

# --- 4. FRONTEND STREAMLIT ---
col_logo, col_titulo = st.columns([1, 8])
with col_logo:
    if os.path.exists("asa_logobrasao_verde.png"): 
        st.image("asa_logobrasao_verde.png")
    else: 
        st.markdown("<h2 style='text-align: center; color: #436468; margin-top: 15px;'>ASA</h2>", unsafe_allow_html=True)
with col_titulo:
    st.title("ASA - Radar Estratégico de Licitações")
    st.markdown("Monitoramento inteligente via **PNCP** e Big Data.")
    
st.divider()

aba_busca, aba_interesse, aba_rastreador, aba_prospeccao, aba_noticias = st.tabs([
    "🔍 Nova Busca", 
    "⭐ Unificador (Pipeline)", 
    "📈 Rastreador", 
    "🎯 Radar de Prospecção", 
    "🌐 Radar de Prévias"
])

# ==========================================
# ABA 1: BUSCA INICIAL SEGURA
# ==========================================
with aba_busca:
    col_filtros, col_resultados = st.columns([1, 3], gap="large")
    
    with col_filtros:
        st.subheader("🎯 Parâmetros Regionais")
        estados_selecionados = st.multiselect("Estados de Interesse (UF):", LISTA_UFS)
        
        st.subheader("📅 Período")
        hoje = datetime.now()
        data_inicio = st.date_input("Data Inicial:", value=hoje - timedelta(days=15))
        data_fim = st.date_input("Data Final:", value=hoje)
        
        st.subheader("🔎 Filtros Avançados")
        # Palavra-chave vem vazia por padrão
        palavras_chave = st.text_input("Palavras-chave (separadas por vírgula):", "")
        modalidades_selecionadas = st.multiselect("Modalidades:", list(MAPA_MODALIDADES.keys()), default=["Concorrência", "Leilão"])
        
        st.subheader("💰 Filtro Financeiro")
        # Valor de volta para 100 Milhões
        valor_min = st.number_input("Valor Mín. (R$):", value=100000000.0, step=1000000.0)
        valor_max = st.number_input("Valor Máx. (R$):", value=5000000000.0, step=1000000.0)
        buscar = st.button("🔍 Mapear Oportunidades")

    with col_resultados:
        if buscar:
            with st.spinner("🔍 Iniciando sistema de varredura..."):
                dados_brutos, erros = buscar_licitacoes_periodo(data_inicio, data_fim, modalidades_selecionadas)
                if erros: 
                    st.warning(f"Alguns blocos apresentaram oscilação no Governo: {', '.join(erros)}")
                
                if dados_brutos:
                    st.session_state['resultados_busca'] = filtrar_dados(dados_brutos, palavras_chave, valor_min, valor_max, estados_selecionados)
                else:
                    st.session_state['resultados_busca'] = pd.DataFrame()
                
                st.session_state['busca_realizada'] = True

        if st.session_state['busca_realizada']:
            df_atual = st.session_state['resultados_busca']
            if not df_atual.empty:
                if "Acompanhar" not in df_atual.columns: 
                    df_atual.insert(0, "Acompanhar", False)
                
                st.write("### 📌 Resultados da Busca")
                m1, m2 = st.columns(2)
                m1.metric("Encontradas", f"{len(df_atual)}")
                m2.metric("Volume", f"R$ {df_atual['Valor Estimado'].sum():,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                
                df_editado = st.data_editor(
                    df_atual, 
                    column_config={
                        "Acompanhar": st.column_config.CheckboxColumn("⭐ Salvar", default=False), 
                        "Link": st.column_config.LinkColumn("Edital", display_text="Acessar")
                    }, 
                    disabled=["Identificação", "Status/Fase", "Dias Restantes", "Data da Sessão", "Última Atualização", "UF", "Órgão", "Modalidade", "Objeto", "Valor Estimado", "Data Publicação", "Anotações Equipe"],
                    hide_index=True, 
                    use_container_width=True
                )
                
                if st.button("💾 Enviar selecionadas para o Pipeline (Aba Unificador)"):
                    selecionadas = df_editado[df_editado["Acompanhar"] == True].drop(columns=["Acompanhar"])
                    if not selecionadas.empty:
                        st.session_state['licitacoes_salvas'] = pd.concat([st.session_state['licitacoes_salvas'], selecionadas]).drop_duplicates(subset=["Identificação"])
                        st.success("Salvas! Vá para a aba Unificador para consolidar com sua planilha antiga.")
                    else:
                        st.warning("Selecione pelo menos uma licitação.")
            else: 
                st.info("Nenhuma licitação encontrada nos filtros estipulados.")

# ==========================================
# ABA 2: UNIFICADOR (O CRM COMPLETO)
# ==========================================
with aba_interesse:
    st.subheader("⭐ Consolidar Pipeline de Vendas")
    st.markdown("Faça upload da sua Planilha Mestre. O robô vai adicionar suas novas seleções sem perder **nenhuma anotação antiga** ou ganhador que já existia.")
    
    arquivo_base = st.file_uploader("📂 Upload do Pipeline Master (.xlsx)", type=["xlsx"], key="up_mestre")
    df_export = st.session_state['licitacoes_salvas'].copy()
    
    if not df_export.empty:
        for col, default_val in [("Empresa Vencedora (Alvo)", "Ainda sem vencedor publicado"), 
                                 ("CNPJ do Alvo", "-"), 
                                 ("Valor Arrematado", 0.0)]:
            if col not in df_export.columns:
                df_export[col] = default_val
    
    if arquivo_base:
        df_antigo = load_master_excel(arquivo_base)
        if not df_antigo.empty:
            df_export = pd.concat([df_antigo, df_export]).drop_duplicates(subset=["Identificação"], keep="first")
            st.success("✅ Base antiga consolidada com as novas buscas!")
            
    if not df_export.empty:
        st.write(f"Total de registros na sua base: **{len(df_export)}**")
        buffer, _, _ = gerar_excel_pipeline(df_export)
        st.download_button(
            "📥 Baixar Pipeline Master Atualizado", 
            buffer.getvalue(), 
            f"Pipeline_ASA_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        if st.button("🗑️ Limpar Sessão Temporária"):
            st.session_state['licitacoes_salvas'] = pd.DataFrame()
            st.rerun()

# ==========================================
# ABA 3: RASTREADOR TURBO
# ==========================================
with aba_rastreador:
    st.subheader("📈 Rastreador de Status")
    st.markdown("O Rastreador **lê todas as abas** do seu Pipeline Master, checa se as datas ou fases mudaram, e devolve a planilha preservando sua estrutura e anotações.")
    
    arquivo_rastreio = st.file_uploader("📂 Upload do Pipeline Master (.xlsx)", type=["xlsx"], key="up_rastreio")
    
    if arquivo_rastreio and st.button("🔄 Rastrear Turbo"):
        df_rastrear = load_master_excel(arquivo_rastreio)
        
        if not df_rastrear.empty and "Link" in df_rastrear.columns:
            progresso = st.progress(0)
            total = len(df_rastrear)
            
            with st.spinner("Atualizando processos..."):
                with ThreadPoolExecutor(max_workers=10) as executor:
                    futuros = [executor.submit(worker_rastrear, idx, row) for idx, row in df_rastrear.iterrows()]
                    for i, futuro in enumerate(as_completed(futuros)):
                        res = futuro.result()
                        idx = res["index"]
                        df_rastrear.at[idx, "Status/Fase"] = res["Status/Fase"]
                        df_rastrear.at[idx, "Última Atualização"] = res["Última Atualização"]
                        df_rastrear.at[idx, "Data da Sessão"] = res["Data da Sessão"]
                        df_rastrear.at[idx, "Dias Restantes"] = res["Dias Restantes"]
                        progresso.progress((i + 1) / total)
                        
            st.success("✅ Rastreamento concluído!")
            buffer, _, _ = gerar_excel_pipeline(df_rastrear)
            st.download_button(
                "📥 Baixar Pipeline Rastreado", 
                buffer.getvalue(), 
                "Pipeline_ASA_Rastreado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else: 
            st.error("Erro ao processar o arquivo. Verifique se as colunas estão corretas.")

# ==========================================
# ABA 4: RADAR DE PROSPECÇÃO (O FUNIL INTELIGENTE)
# ==========================================
with aba_prospeccao:
    st.subheader("🎯 Radar de Prospecção (CRM)")
    st.markdown("O Radar processa **apenas as licitações 'Em Trânsito'**. Se ele achar novos vencedores, os transfere de aba automaticamente.")
    
    arquivo_prospeccao = st.file_uploader("📂 Upload do Pipeline Master (.xlsx)", type=["xlsx"], key="up_prospeccao")
    
    if arquivo_prospeccao and st.button("🔎 Gerar Leads Turbo"):
        df_prosp = load_master_excel(arquivo_prospeccao)
        
        if not df_prosp.empty and "Link" in df_prosp.columns:
            mask_transito = df_prosp["Empresa Vencedora (Alvo)"] == "Ainda sem vencedor publicado"
            df_to_process = df_prosp[mask_transito].copy()
            df_already_won = df_prosp[~mask_transito].copy()
            
            progresso = st.progress(0)
            total = len(df_to_process)
            alvos = []
            
            if total > 0:
                with st.spinner(f"Investigando {total} contratos em andamento no Governo..."):
                    with ThreadPoolExecutor(max_workers=10) as executor:
                        futuros = [executor.submit(worker_prospeccao, row) for _, row in df_to_process.iterrows()]
                        for i, futuro in enumerate(as_completed(futuros)):
                            alvos.append(futuro.result())
                            progresso.progress((i + 1) / total)
                            
            df_processados = pd.DataFrame(alvos) if alvos else pd.DataFrame(columns=df_prosp.columns)
            
            df_final_completo = pd.concat([df_already_won, df_processados], ignore_index=True)
            
            buffer, count_transito, count_vencedores = gerar_excel_pipeline(df_final_completo)
            
            st.success(f"🎯 Concluído! Funil organizado: {count_transito} processos Em Trânsito e {count_vencedores} Vencedores Revelados.")
            
            st.download_button(
                "📥 Baixar Pipeline Master (Com Abas de Vendas)", 
                buffer.getvalue(), 
                f"Pipeline_Prospeccao_ASA_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else: 
            st.error("Erro ao ler o arquivo. A coluna 'Link' não foi encontrada.")

# ==========================================
# ABA 5: CLIPPING DE NOTÍCIAS INTELIGENTE
# ==========================================
with aba_noticias:
    st.subheader("🌐 Radar de Mercado Inteligente (Filtro Anti-Lixo Ativo)")
    st.markdown("O robô vasculha notícias e aplica uma **filtragem semântica local** para focar em infraestrutura.")
    
    lista_fases = ['Consulta Pública', 'Audiência Pública', 'Manifestação de Interesse', 'Aviso de Licitação', 'Estudos Técnicos Preliminares']
    lista_temas = ['Concessão', 'PPP', 'Saneamento', 'Rodovia', 'Pedágio', 'Iluminação Pública', 'Privatização']
    
    col1, col2, col3 = st.columns(3)
    fases_alvo = col1.multiselect("Fases/Atos:", lista_fases, default=lista_fases)
    temas_alvo = col2.multiselect("Setores/Temas:", lista_temas, default=['Concessão', 'PPP', 'Saneamento', 'Rodovia', 'Privatização'])
    filtro_tempo = col3.selectbox("Tempo de Busca:", [("Últimas 24h", 1), ("Última Semana", 7), ("Último Mês", 30)], format_func=lambda x: x[0])
    
    arquivo_clipping = st.file_uploader("📂 Opcional: Upload do Clipping Antigo (.xlsx)", type=["xlsx"])
        
    if st.button("📰 Iniciar Varredura Inteligente"):
        if fases_alvo and temas_alvo:
            total_buscas = len(fases_alvo) * len(temas_alvo)
            progresso_news = st.progress(0)
            dfs_noticias = []
            
            with st.spinner(f"Disparando {total_buscas} buscas cruzadas com filtro anti-lixo..."):
                with ThreadPoolExecutor(max_workers=5) as executor:
                    futuros = [
                        executor.submit(worker_rss, fase, tema, filtro_tempo[1]) 
                        for fase in fases_alvo for tema in temas_alvo
                    ]
                    for i, futuro in enumerate(as_completed(futuros)):
                        resultado_df = futuro.result()
                        if not resultado_df.empty: 
                            dfs_noticias.append(resultado_df)
                        progresso_news.progress((i + 1) / total_buscas)
                        
            if dfs_noticias:
                df_novas = pd.concat(dfs_noticias).drop_duplicates(subset=["Link da Notícia"])
                df_final = df_novas.copy()
                
                if arquivo_clipping:
                    try:
                        df_antigo = pd.read_excel(arquivo_clipping)
                        df_final = pd.concat([df_antigo, df_novas]).drop_duplicates(subset=["Link da Notícia"], keep="first")
                        st.success("✅ Base unida preservando anotações antigas!")
                    except: 
                        st.error("Erro ao ler a planilha de clipping antiga.")
                        
                st.success(f"✅ {len(df_final)} notícias valiosas consolidadas!")
                st.dataframe(df_final, hide_index=True, use_container_width=True)
                
                buffer_news = io.BytesIO()
                with pd.ExcelWriter(buffer_news, engine='openpyxl') as writer:
                    df_final.to_excel(writer, index=False, sheet_name='Clipping Estratégico ASA')
                    aplicar_estilo_excel(writer, df_final, 'Clipping Estratégico ASA')
                    
                st.download_button(
                    label="📥 Baixar Clipping Limpo Excel (.xlsx)",
                    data=buffer_news.getvalue(),
                    file_name=f"Clipping_ASA_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else: 
                st.warning("Nenhuma notícia de infraestrutura/licitação encontrada para essas combinações no período selecionado.")
        else: 
            st.warning("Selecione pelo menos uma Fase e um Tema.")
